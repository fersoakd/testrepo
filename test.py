from openpyxl import Workbook
from datetime import datetime

#Diccionario de ejemplo
datos = {'01/01/2016':100, '02/01/2016':102, '03/01/2016':103, '04/01/2016':104,
         '05/01/2016':105, '06/01/2016':106, '07/01/2016':107, '08/01/2016':108,
         '09/01/2016':109, '10/01/2016':110, '11/01/2016':111, '12/01/2016':112}


wb = Workbook()
ruta = 'salida.xlsx'

hoja = wb.active
hoja.title = "Fecha-Valor"

fila = 1 #Fila donde empezamos
col_fecha = 1 #Columna donde guardamos las fechas
col_dato = 2 #Columna donde guardamos el dato asociados a cada fecha

for fecha, dato in datos.items():
    hoja.cell(column=col_fecha, row=fila, value=datetime.strptime(fecha, '%d/%m/%Y').date())
    hoja.cell(column=col_dato, row=fila, value=dato)
    fila+=1

wb.save(filename = ruta)